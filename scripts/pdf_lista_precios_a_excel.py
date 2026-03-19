from __future__ import annotations

import json
import os
import re
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from html import escape
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PDF_TEXT = Path("PDFs/co_lista_precios_extract.txt")
OUTPUT_XLSX = Path("PDFs/CO-lista-de-precios-detallada-2025.xlsx")
OUTPUT_HTML = Path("lista-precios-detallada-2025.html")
OUTPUT_PRICE_PDF = Path("PDFs/CO-lista-de-precios-detallada-2025.pdf")
OUTPUT_SALES_PLAN_PDF = Path("PDFs/sales-performance-plan-overview-infographic-latam-2025.pdf")
IMAGE_CACHE_JSON = Path("PDFs/CO-lista-de-precios-detallada-2025-images.json")
CATEGORY_JSON = Path("nuskin-colombia-productos.json")
FALLBACK_PRODUCT_URL = "https://www.nuskin.com/content/nuskin/es_CO/products/product.{sku}.html"
GRAPHQL_ENDPOINT = "https://apis.nuskin.com/product/graphql"
GRAPHQL_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-CO,es;q=0.9,en;q=0.8",
}

COLUMNS = [
    "Categoria",
    "SKU",
    "Descripcion del producto",
    "Precio Publico",
    "Precio Publico sin Impuesto",
    "Precio Miembro",
    "Precio Miembro sin Impuesto",
    "Bono minorista sin Impuesto",
    "VCV",
    "VV",
    "4%",
    "8%",
    "12%",
    "16%",
    "20%",
    "24%",
]

PRICE_COLUMNS = COLUMNS[3:]
BONUS_COLUMNS = ["4%", "8%", "12%", "16%", "20%", "24%"]
DETAIL_FILE_RE = re.compile(r"^\d{8}(?:_[A-Za-z0-9_]+)?$")

HEADER_PREFIXES = (
    "SKU Descrip",
    "4% 8% 12% 16% 20% 24%",
    "m 4% 8% 12% 16% 20% 24%",
    "Bonos por Compartir",
    "WellSpa iO",
    "Kits de Inicio",
    "LumiSpa iO Kits",
    "Limpiadores y Tramientos LumiSpa",
    "Cabezales y accesorios LumiSpa iO",
    "Epoch",
    "Nu Colour",
    "Sunright",
    "Nutricionales",
    "Soluciones",
    "Nu Skin 180",
    "Pharmanex",
    "Cuidado Facial",
    "Cuidado Dental",
    "Tratamientos ageLOC",
    "ageLOC Boost",
    "ageLOC Galvanic Spa",
    "Suscripción",
    "HORARIO DE",
    "Lunes a Viernes",
    "Sábados",
    "NOTAS",
    "Órdenes",
    "a) VCV",
    "b) Tu Bono",
    "*(excluyendo",
    "c) Esta lista",
    "para ver y estar",
    "1",
    "2",
    "3",
)

NUMBER_PATTERN = r"\(?-?[\d,]+(?:\.\d+)?\)?|-"
ROW_REGEX = re.compile(
    rf"^(?P<sku>\d{{8}})\s+(?P<descripcion>.*?)\s+"
    rf"(?P<n1>{NUMBER_PATTERN})\s+(?P<n2>{NUMBER_PATTERN})\s+"
    rf"(?P<n3>{NUMBER_PATTERN})\s+(?P<n4>{NUMBER_PATTERN})\s+"
    rf"(?P<n5>{NUMBER_PATTERN})\s+(?P<n6>{NUMBER_PATTERN})\s+"
    rf"(?P<n7>{NUMBER_PATTERN})\s+(?P<n8>{NUMBER_PATTERN})\s+"
    rf"(?P<n9>{NUMBER_PATTERN})\s+(?P<n10>{NUMBER_PATTERN})\s+"
    rf"(?P<n11>{NUMBER_PATTERN})\s+(?P<n12>{NUMBER_PATTERN})\s+"
    rf"(?P<n13>{NUMBER_PATTERN})$"
)


def clean_text(value: str) -> str:
    replacements = {
        "Â®": "®",
        "Ž": "®",
        "â€”": "-",
        "â€“": "-",
        "â€¯": " ",
        "Ëš": "°",
        "Â°": "°",
        "Ã¡": "á",
        "Ã©": "é",
        "Ã­": "í",
        "Ã³": "ó",
        "Ãº": "ú",
        "Ã±": "ñ",
        "Ã“": "Ó",
        "Ã‰": "É",
        "Ãš": "Ú",
        "Ã‘": "Ñ",
        "Â": "",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalize_key(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(char for char in value if not unicodedata.combining(char))
    return value.lower()


def parse_number(value: str) -> float | None:
    value = value.strip()
    if value == "-":
        return None
    negative = value.startswith("(") and value.endswith(")")
    value = value.strip("()").replace(",", "")
    number = float(value)
    return -number if negative else number


def load_product_metadata() -> dict[str, dict[str, str]]:
    if not CATEGORY_JSON.exists():
        return {}

    payload = json.loads(CATEGORY_JSON.read_text(encoding="utf-8-sig"))
    products = payload.get("products", [])

    metadata: dict[str, dict[str, str]] = {}
    for product in products:
        sku = str(product.get("ProductId") or "").strip()
        category = str(product.get("Category") or "").strip()
        link = str(product.get("Link") or "").strip()
        if not sku or not category:
            continue
        metadata[sku] = {
            "name": clean_text(str(product.get("Name") or "")),
            "description": clean_text(str(product.get("Description") or "")),
            "title": "",
            "category": category.split(" / ")[-1].strip(),
            "link": link,
            "image_url": str(product.get("ImageUrl") or "").strip(),
            "thumbnail_url": str(product.get("ThumbnailUrl") or "").strip(),
            "has_image": str(bool(product.get("HasOfficialImage"))).lower(),
        }

    return metadata


def load_image_cache() -> dict[str, dict[str, str]]:
    if not IMAGE_CACHE_JSON.exists():
        return {}

    payload = json.loads(IMAGE_CACHE_JSON.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        return {
            str(sku): {
                "title": str(item.get("title") or "").strip(),
                "image_url": str(item.get("image_url") or "").strip(),
                "thumbnail_url": str(item.get("thumbnail_url") or "").strip(),
                "has_image": str(bool(item.get("has_image"))).lower(),
            }
            for sku, item in payload.items()
        }
    return {}


def save_image_cache(cache: dict[str, dict[str, str]]) -> None:
    IMAGE_CACHE_JSON.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def pick_graphql_image(product: dict[str, object], sku: str) -> dict[str, str]:
    variants = product.get("variants") or []
    for variant in variants:
        if str(variant.get("sku") or "").strip() != sku:
            continue
        images = [image for image in variant.get("productImages") or [] if image.get("url")]
        if images:
            return images[0]

    for image in product.get("productImages") or []:
        if image.get("url"):
            return image

    for variant in variants:
        images = [image for image in variant.get("productImages") or [] if image.get("url")]
        if images:
            return images[0]

    return {}


def fetch_graphql_image_record(sku: str) -> dict[str, str]:
    query = f"""
query {{
  productById(id:"{sku}", market:"CO", language:"es") {{
    title
    productImages {{ url thumbnail alt }}
    variants {{
      sku
      productImages {{ url thumbnail alt }}
    }}
  }}
}}
"""
    try:
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": query},
            headers=GRAPHQL_HEADERS,
            timeout=40,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return {
            "title": "",
            "image_url": "",
            "thumbnail_url": "",
            "has_image": "false",
        }

    product = (payload.get("data") or {}).get("productById") or {}
    image = pick_graphql_image(product, sku)
    image_url = str(image.get("url") or "").strip()
    thumbnail_url = str(image.get("thumbnail") or image_url).strip()

    return {
        "title": clean_text(str(product.get("title") or "")),
        "image_url": image_url,
        "thumbnail_url": thumbnail_url,
        "has_image": str(bool(image_url)).lower(),
    }


def enrich_product_metadata_with_graphql(
    product_metadata: dict[str, dict[str, str]],
    skus: list[str],
) -> dict[str, dict[str, str]]:
    cache = load_image_cache()
    missing = [sku for sku in skus if sku and sku not in cache]

    if missing:
        with ThreadPoolExecutor(max_workers=8) as executor:
            future_map = {
                executor.submit(fetch_graphql_image_record, sku): sku
                for sku in missing
            }
            for future in as_completed(future_map):
                sku = future_map[future]
                cache[sku] = future.result()

        save_image_cache(cache)

    for sku in skus:
        metadata = product_metadata.setdefault(
            sku,
            {
                "name": "",
                "description": "",
                "title": "",
                "category": "",
                "link": FALLBACK_PRODUCT_URL.format(sku=sku),
                "image_url": "",
                "thumbnail_url": "",
                "has_image": "false",
            },
        )
        cached = cache.get(sku, {})
        if cached.get("title"):
            metadata["title"] = cached["title"]
        if cached.get("image_url"):
            metadata["image_url"] = cached["image_url"]
        if cached.get("thumbnail_url"):
            metadata["thumbnail_url"] = cached["thumbnail_url"]
        if cached.get("has_image"):
            metadata["has_image"] = cached["has_image"]

    return product_metadata


def infer_category(sku: str, description: str, product_metadata: dict[str, dict[str, str]]) -> str:
    if sku in product_metadata:
        return product_metadata[sku]["category"]

    text = normalize_key(description)
    rules = [
        (("ap 24", "toothpaste"), "AP-24"),
        (("epoch", "cica balm", "polishing bar"), "Epoch"),
        (("nu colour", "lash + brow", "peptide pout"), "Nu Colour"),
        (("sunright", "solar screen"), "Sunright"),
        (("wellspa", "body io"), "WellSpa iO"),
        (("lumispa", "boost", "galvanic", "tru face", "ageloc"), "ageLOC"),
        (("g3", "omega", "collagen", "vitameal"), "Pharmanex"),
        (("face wash", "180"), "Nu Skin 180"),
    ]

    for keywords, category in rules:
        if any(keyword in text for keyword in keywords):
            return category

    return "Otros"


def extract_rows(text: str, product_metadata: dict[str, dict[str, str]]) -> list[dict[str, object]]:
    lines = [clean_text(re.sub(r"\s+", " ", line).strip()) for line in text.splitlines()]
    lines = [line for line in lines if line]

    rows: list[dict[str, object]] = []
    current: str | None = None

    def flush() -> None:
      nonlocal current
      if not current:
          return
      match = ROW_REGEX.match(current)
      if match:
          groups = match.groupdict()
          rows.append(
              {
                  "Categoria": infer_category(groups["sku"], groups["descripcion"], product_metadata),
                  "SKU": groups["sku"],
                  "Descripcion del producto": clean_text(groups["descripcion"]),
                  "Precio Publico": parse_number(groups["n1"]),
                  "Precio Publico sin Impuesto": parse_number(groups["n2"]),
                  "Precio Miembro": parse_number(groups["n3"]),
                  "Precio Miembro sin Impuesto": parse_number(groups["n4"]),
                  "Bono minorista sin Impuesto": parse_number(groups["n5"]),
                  "VCV": parse_number(groups["n6"]),
                  "VV": parse_number(groups["n7"]),
                  "4%": parse_number(groups["n8"]),
                  "8%": parse_number(groups["n9"]),
                  "12%": parse_number(groups["n10"]),
                  "16%": parse_number(groups["n11"]),
                  "20%": parse_number(groups["n12"]),
                  "24%": parse_number(groups["n13"]),
              }
          )
      current = None

    for line in lines:
        if line.startswith(HEADER_PREFIXES) or re.fullmatch(r"\d+", line):
            flush()
            continue

        if re.match(r"^\d{8}\b", line):
            flush()
            current = line
            continue

        if current:
            current = f"{current} {line}"

    flush()
    return rows


def autosize_sheet(ws) -> None:
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 42)


def color_to_css(color) -> str:
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return ""
    value = str(rgb)
    if len(value) == 8:
        value = value[2:]
    if len(value) != 6:
        return ""
    return f"#{value}"


def format_html_value(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"{value:,.1f}"
    return str(value)


def get_cell_fill_color(cell) -> str:
    if cell.fill.fill_type != "solid":
        return ""
    return color_to_css(cell.fill.fgColor)


def render_cell_content(cell) -> str:
    content = escape(format_html_value(cell))
    if cell.hyperlink and cell.hyperlink.target:
        link = escape(cell.hyperlink.target, quote=True)
        return f'<a href="{link}" target="_blank" rel="noreferrer">{content}</a>'
    return content


def build_price_payload(ws, row_idx: int) -> str:
    payload: dict[str, str] = {}
    for col_idx in range(4, ws.max_column + 1):
        label = str(ws.cell(row=1, column=col_idx).value or "").strip()
        payload[label] = format_html_value(ws.cell(row=row_idx, column=col_idx))
    return escape(json.dumps(payload, ensure_ascii=False), quote=True)


def build_image_button(
    sku: str,
    product_name: str,
    product_metadata: dict[str, dict[str, str]],
    class_name: str = "img-emoji",
) -> str:
    product = product_metadata.get(sku, {})
    image_url = product.get("image_url", "") or product.get("thumbnail_url", "")
    title = "Ver imagen oficial" if image_url else "Ver vista referencial"
    state = "has-image" if image_url else "is-placeholder"
    return (
        f'<button type="button" class="{class_name} {state}" '
        f'data-sku="{escape(sku, quote=True)}" '
        f'data-name="{escape(product_name, quote=True)}" '
        f'data-image="{escape(image_url, quote=True)}" '
        f'title="{title}" aria-label="{title}">🖼️</button>'
    )


def build_price_button(
    sku: str,
    product_name: str,
    price_payload: str,
    class_name: str = "price-emoji",
) -> str:
    title = "Ver precios de la fila"
    return (
        f'<button type="button" class="{class_name}" '
        f'data-sku="{escape(sku, quote=True)}" '
        f'data-name="{escape(product_name, quote=True)}" '
        f'data-prices="{price_payload}" '
        f'title="{title}" aria-label="{title}">💲</button>'
    )


def build_prompt_button(
    sku: str,
    product_name: str,
    product_metadata: dict[str, dict[str, str]],
    class_name: str = "prompt-emoji",
) -> str:
    product = product_metadata.get(sku, {})
    official_url = (product.get("link") or FALLBACK_PRODUCT_URL.format(sku=sku)).strip()
    image_url = (product.get("image_url") or product.get("thumbnail_url") or "").strip()
    category = str(product.get("category", "") or "").strip()
    title = "Copiar prompt para imagen de WhatsApp"
    return (
        f'<button type="button" class="{class_name}" '
        f'data-sku="{escape(sku, quote=True)}" '
        f'data-name="{escape(product_name, quote=True)}" '
        f'data-url="{escape(official_url, quote=True)}" '
        f'data-category="{escape(category, quote=True)}" '
        f'data-image="{escape(image_url, quote=True)}" '
        f'title="{title}" aria-label="{title}">📋</button>'
    )


def detail_page_href(sku: str, product_metadata: dict[str, dict[str, str]]) -> str:
    sku = str(sku or "").strip()
    if not sku:
        return ""
    return str(product_metadata.get(sku, {}).get("detail_page", "") or "").strip()


def build_detail_link(
    sku: str,
    product_metadata: dict[str, dict[str, str]],
    class_name: str = "detail-emoji",
) -> str:
    href = detail_page_href(sku, product_metadata)
    if not href:
        return ""
    title = "Abrir ficha detallada"
    return (
        f'<a class="{class_name}" href="{escape(href, quote=True)}" '
        f'title="{title}" aria-label="{title}">📄</a>'
    )


def build_action_group(
    sku: str,
    product_name: str,
    product_metadata: dict[str, dict[str, str]],
    price_payload: str,
) -> str:
    actions = [
        build_image_button(sku, product_name, product_metadata),
        build_price_button(sku, product_name, price_payload),
        build_prompt_button(sku, product_name, product_metadata),
        build_detail_link(sku, product_metadata),
    ]
    return f'<span class="action-group">{"".join(action for action in actions if action)}</span>'


def render_description_with_actions(
    cell,
    sku: str,
    product_metadata: dict[str, dict[str, str]],
    price_payload: str,
) -> str:
    return (
        '<div class="desc-inline">'
        f"{render_cell_content(cell)}"
        f"{build_action_group(sku, format_html_value(cell), product_metadata, price_payload)}"
        "</div>"
    )


def build_html_table(ws, product_metadata: dict[str, dict[str, str]]) -> str:
    merged_starts: dict[tuple[int, int], tuple[int, int]] = {}
    merged_skips: set[tuple[int, int]] = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        merged_starts[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                if (row_idx, col_idx) != (min_row, min_col):
                    merged_skips.add((row_idx, col_idx))

    colgroup = []
    for column_idx in range(1, ws.max_column + 1):
        width = ws.column_dimensions[get_column_letter(column_idx)].width or 14
        colgroup.append(f'<col style="width:{int(width * 8)}px">')

    rows_html: list[str] = ['<div class="table-wrap desktop-view">', '<table class="price-table">', "<colgroup>"]
    rows_html.extend(colgroup)
    rows_html.append("</colgroup>")

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        tag = "th" if row_idx == 1 else "td"
        row_cells: list[str] = []
        for cell in row:
            coord = (cell.row, cell.column)
            if coord in merged_skips:
                continue

            attrs: list[str] = []
            if coord in merged_starts:
                rowspan, colspan = merged_starts[coord]
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')

            styles: list[str] = []
            fill_color = get_cell_fill_color(cell)
            if fill_color:
                styles.append(f"background:{fill_color}")

            font_color = ""
            if getattr(cell.font.color, "type", "") == "rgb":
                font_color = color_to_css(cell.font.color)
            if font_color and font_color != "#000000":
                styles.append(f"color:{font_color}")

            horizontal = cell.alignment.horizontal
            if horizontal:
                styles.append(f"text-align:{horizontal}")
            elif row_idx > 1 and isinstance(cell.value, (int, float)):
                styles.append("text-align:right")

            if cell.font.bold:
                styles.append("font-weight:700")

            if cell.alignment.wrap_text:
                styles.append("white-space:normal")

            if cell.column == 1 and row_idx > 1:
                styles.append("vertical-align:middle")

            if styles:
                attrs.append(f'style="{"; ".join(styles)}"')

            classes: list[str] = []
            if row_idx > 1 and isinstance(cell.value, (int, float)):
                classes.append("num")
            if cell.column == 1 and row_idx > 1:
                classes.append("cat")
            if classes:
                attrs.append(f'class="{" ".join(classes)}"')

            if row_idx > 1 and cell.column == 3:
                sku = str(ws.cell(row=row_idx, column=2).value or "").strip()
                price_payload = build_price_payload(ws, row_idx)
                content = render_description_with_actions(cell, sku, product_metadata, price_payload)
            else:
                content = render_cell_content(cell)
            row_cells.append(f"<{tag} {' '.join(attrs)}>{content}</{tag}>")

        rows_html.append(f"<tr>{''.join(row_cells)}</tr>")

    rows_html.append("</table>")
    rows_html.append("</div>")
    return "\n".join(rows_html)


def build_mobile_cards(ws, product_metadata: dict[str, dict[str, str]]) -> str:
    headers = [str(cell.value or "") for cell in ws[1]]
    sections: list[str] = ['<div class="mobile-sheet">']

    current_category = ""
    current_color = "#607D8B"
    section_open = False

    for row_idx in range(2, ws.max_row + 1):
        category_cell = ws.cell(row=row_idx, column=1)
        if category_cell.value:
            if section_open:
                sections.append("</div></section>")
            current_category = str(category_cell.value)
            current_color = get_cell_fill_color(category_cell) or "#607D8B"
            sections.append(
                "\n".join(
                    [
                        f'<section class="mobile-category" style="--category-color:{current_color}">',
                        f'<header class="mobile-category-title">{escape(current_category)}</header>',
                        '<div class="mobile-card-list">',
                    ]
                )
            )
            section_open = True

        description_cell = ws.cell(row=row_idx, column=3)
        sku_cell = ws.cell(row=row_idx, column=2)
        price_payload = build_price_payload(ws, row_idx)
        metrics: list[str] = []

        for col_idx in range(4, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            metrics.append(
                "\n".join(
                    [
                        '<div class="metric-item">',
                        f'<span class="metric-label">{escape(headers[col_idx - 1])}</span>',
                        f'<strong class="metric-value">{escape(format_html_value(cell))}</strong>',
                        "</div>",
                    ]
                )
            )

        sections.append(
            "\n".join(
                [
                    '<article class="mobile-card">',
                    '<div class="mobile-card-top">',
                    f'<span class="sku-chip">SKU {escape(format_html_value(sku_cell))}</span>',
                    '<div class="product-title-row">',
                    f'<h3 class="product-title">{render_cell_content(description_cell)}</h3>',
                    build_action_group(
                        str(sku_cell.value or "").strip(),
                        format_html_value(description_cell),
                        product_metadata,
                        price_payload,
                    ),
                    "</div>",
                    "</div>",
                    '<div class="metric-grid">',
                    "".join(metrics),
                    "</div>",
                    "</article>",
                ]
            )
        )

    if section_open:
        sections.append("</div></section>")

    sections.append("</div>")
    return "\n".join(sections)


def export_workbook_html(
    workbook,
    output_path: Path,
    product_metadata: dict[str, dict[str, str]],
) -> None:
    sections = []
    tab_buttons = []

    for index, sheet_name in enumerate(workbook.sheetnames):
        ws = workbook[sheet_name]
        active = " is-active" if index == 0 else ""
        safe_id = f"sheet-{index + 1}"
        tab_buttons.append(
            f'<button class="tab{active}" data-target="{safe_id}" type="button">{escape(sheet_name)}</button>'
        )
        sections.append(
            "\n".join(
                [
                    f'<section id="{safe_id}" class="sheet-panel{active}">',
                    f"<h2>{escape(sheet_name)}</h2>",
                    build_html_table(ws, product_metadata),
                    build_mobile_cards(ws, product_metadata),
                    "</section>",
                ]
            )
        )

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Lista de Precios Nu Skin Colombia 2025</title>
  <style>
    :root {{
      --bg: #f4f7fb;
      --card: #ffffff;
      --text: #233142;
      --muted: #5f6b7a;
      --line: #d7dee7;
      --shadow: 0 18px 48px rgba(25, 52, 80, 0.12);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: 'Segoe UI', Arial, sans-serif;
      background:
        radial-gradient(circle at top right, rgba(21, 101, 192, 0.08), transparent 28%),
        linear-gradient(180deg, #f8fbff 0%, var(--bg) 100%);
      color: var(--text);
    }}
    .page {{
      max-width: 1600px;
      margin: 0 auto;
      padding: 28px 18px 40px;
    }}
    .hero {{
      background: var(--card);
      border: 1px solid rgba(21, 101, 192, 0.14);
      border-radius: 20px;
      box-shadow: var(--shadow);
      padding: 24px 26px;
      margin-bottom: 18px;
    }}
    h1 {{
      margin: 0 0 8px;
      color: #1565c0;
      font-size: clamp(1.8rem, 3vw, 2.6rem);
    }}
    .hero p {{
      margin: 0;
      color: var(--muted);
      line-height: 1.55;
      max-width: 900px;
    }}
    .prompt-panel {{
      margin-top: 18px;
      padding-top: 18px;
      border-top: 1px solid #e3ebf4;
      display: grid;
      gap: 12px;
    }}
    .prompt-panel-title {{
      margin: 0;
      color: #233142;
      font-size: 1.05rem;
    }}
    .prompt-help {{
      margin: 0;
      color: var(--muted);
      line-height: 1.5;
      max-width: none;
    }}
    .contact-grid {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }}
    .field {{
      display: grid;
      gap: 6px;
    }}
    .field span {{
      color: #1565c0;
      font-size: 0.82rem;
      font-weight: 800;
      letter-spacing: 0.02em;
    }}
    .field input {{
      width: 100%;
      border: 1px solid #c9d8ea;
      border-radius: 12px;
      padding: 11px 12px;
      font: inherit;
      color: #233142;
      background: #fff;
    }}
    .field input:focus {{
      outline: 2px solid rgba(21, 101, 192, 0.18);
      border-color: #1565c0;
    }}
    .clipboard-feedback {{
      min-height: 1.35rem;
      margin: 0;
      color: #166534;
      font-size: 0.92rem;
      font-weight: 700;
    }}
    .tabs {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin: 0 0 16px;
    }}
    .tab {{
      border: 1px solid #c9d8ea;
      background: #fff;
      color: #1565c0;
      padding: 10px 16px;
      border-radius: 999px;
      font-size: 0.98rem;
      font-weight: 700;
      cursor: pointer;
      transition: 0.2s ease;
    }}
    .tab:hover {{
      background: #edf5ff;
    }}
    .tab.is-active {{
      background: #1565c0;
      color: #fff;
      border-color: #1565c0;
    }}
    .sheet-panel {{
      display: none;
      background: var(--card);
      border: 1px solid rgba(35, 49, 66, 0.08);
      border-radius: 20px;
      box-shadow: var(--shadow);
      padding: 20px;
    }}
    .sheet-panel.is-active {{
      display: block;
    }}
    h2 {{
      margin: 0 0 14px;
      color: #2e7d32;
      font-size: 1.4rem;
    }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 16px;
      background: #fff;
    }}
    .mobile-sheet {{
      display: none;
    }}
    .price-table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 1450px;
      table-layout: fixed;
    }}
    .price-table th,
    .price-table td {{
      border: 1px solid var(--line);
      padding: 10px 12px;
      font-size: 0.95rem;
      line-height: 1.35;
      vertical-align: top;
      white-space: nowrap;
    }}
    .price-table th {{
      position: sticky;
      top: 0;
      z-index: 3;
    }}
    .price-table td.cat {{
      min-width: 130px;
    }}
    .price-table td.num {{
      font-variant-numeric: tabular-nums;
    }}
    .desc-inline {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      max-width: 100%;
      white-space: normal;
    }}
    .action-group {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
      flex: 0 0 auto;
    }}
    .img-emoji,
    .price-emoji,
    .prompt-emoji,
    .detail-emoji {{
      border: 0;
      width: 32px;
      height: 32px;
      border-radius: 999px;
      background: #fff3e0;
      cursor: pointer;
      font-size: 1rem;
      line-height: 1;
      flex: 0 0 auto;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      box-shadow: 0 6px 14px rgba(239, 108, 0, 0.16);
      transition: transform 0.18s ease, background 0.18s ease;
    }}
    .img-emoji:hover,
    .price-emoji:hover,
    .prompt-emoji:hover,
    .detail-emoji:hover {{
      background: #ffe0b2;
      transform: translateY(-1px);
      text-decoration: none;
    }}
    .img-emoji.is-placeholder {{
      background: #eceff1;
      box-shadow: 0 6px 14px rgba(96, 125, 139, 0.12);
    }}
    .price-emoji {{
      background: #eef6ff;
      box-shadow: 0 6px 14px rgba(21, 101, 192, 0.14);
    }}
    .price-emoji:hover {{
      background: #dbeafe;
    }}
    .prompt-emoji {{
      background: #fef3c7;
      box-shadow: 0 6px 14px rgba(217, 119, 6, 0.14);
    }}
    .prompt-emoji:hover {{
      background: #fde68a;
    }}
    .detail-emoji {{
      background: #ecfdf5;
      color: #166534;
      box-shadow: 0 6px 14px rgba(22, 101, 52, 0.12);
    }}
    .detail-emoji:hover {{
      background: #d1fae5;
    }}
    .price-table a {{
      color: #0563c1;
      text-decoration: underline;
      font-weight: 600;
    }}
    .price-table tr:hover td:not(.cat) {{
      background: #f7fbff;
    }}
    .mobile-category {{
      margin-top: 16px;
    }}
    .mobile-category-title {{
      margin: 0 0 12px;
      padding: 12px 14px;
      border-radius: 14px;
      background: linear-gradient(135deg, var(--category-color), #ffffff);
      color: #ffffff;
      font-size: 1rem;
      font-weight: 800;
      letter-spacing: 0.02em;
      text-shadow: 0 1px 2px rgba(0, 0, 0, 0.15);
    }}
    .mobile-card-list {{
      display: grid;
      gap: 12px;
    }}
    .mobile-card {{
      border: 1px solid #dde6f0;
      border-radius: 18px;
      padding: 14px;
      background: #ffffff;
      box-shadow: 0 10px 24px rgba(21, 101, 192, 0.08);
    }}
    .mobile-card-top {{
      display: grid;
      gap: 10px;
      margin-bottom: 12px;
    }}
    .sku-chip {{
      display: inline-flex;
      width: fit-content;
      padding: 6px 10px;
      border-radius: 999px;
      background: #edf5ff;
      color: #1565c0;
      font-size: 0.78rem;
      font-weight: 800;
      letter-spacing: 0.04em;
    }}
    .product-title {{
      margin: 0;
      font-size: 1rem;
      line-height: 1.45;
      color: #233142;
    }}
    .product-title-row {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 10px;
      align-items: start;
    }}
    .product-title a {{
      color: inherit;
      text-decoration: none;
    }}
    .metric-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 10px;
    }}
    .metric-item {{
      border: 1px solid #e3ebf4;
      border-radius: 14px;
      background: #f8fbff;
      padding: 10px 12px;
      min-width: 0;
    }}
    .metric-label {{
      display: block;
      margin-bottom: 4px;
      color: var(--muted);
      font-size: 0.72rem;
      line-height: 1.35;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    .metric-value {{
      display: block;
      color: #233142;
      font-size: 0.98rem;
      line-height: 1.35;
      word-break: break-word;
      font-variant-numeric: tabular-nums;
    }}
    .image-modal[hidden] {{
      display: none !important;
    }}
    .image-modal {{
      position: fixed;
      inset: 0;
      z-index: 1000;
      background: rgba(15, 23, 42, 0.72);
      backdrop-filter: blur(4px);
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 20px;
    }}
    .modal-card {{
      width: min(920px, 100%);
      background: #ffffff;
      border-radius: 22px;
      box-shadow: 0 24px 60px rgba(15, 23, 42, 0.28);
      overflow: hidden;
    }}
    .modal-head {{
      display: flex;
      align-items: start;
      justify-content: space-between;
      gap: 16px;
      padding: 18px 20px 14px;
      border-bottom: 1px solid #e7eef5;
    }}
    .modal-title {{
      margin: 0;
      font-size: 1.15rem;
      line-height: 1.4;
      color: #233142;
    }}
    .modal-subtitle {{
      margin: 4px 0 0;
      color: var(--muted);
      font-size: 0.92rem;
    }}
    .modal-close {{
      border: 0;
      width: 38px;
      height: 38px;
      border-radius: 999px;
      background: #eef4fb;
      color: #233142;
      cursor: pointer;
      font-size: 1.3rem;
      line-height: 1;
      flex: 0 0 auto;
    }}
    .modal-close:hover {{
      background: #dfeaf8;
    }}
    .modal-media {{
      background: linear-gradient(180deg, #f9fbfe 0%, #eef4fb 100%);
      min-height: 320px;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 18px;
    }}
    .modal-image {{
      display: block;
      width: 100%;
      height: auto;
      max-height: 72vh;
      object-fit: contain;
      border-radius: 16px;
      background: #ffffff;
    }}
    .modal-card.compact {{
      width: min(640px, 100%);
    }}
    .price-body {{
      padding: 18px 20px 20px;
      display: grid;
      gap: 10px;
      background: #ffffff;
    }}
    .price-row {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 12px;
      align-items: center;
      border: 1px solid #e3ebf4;
      border-radius: 12px;
      background: #f8fbff;
      padding: 10px 12px;
    }}
    .price-row span {{
      color: var(--muted);
      line-height: 1.4;
    }}
    .price-row strong {{
      color: #233142;
      font-variant-numeric: tabular-nums;
      text-align: right;
    }}
    .footer {{
      color: var(--muted);
      font-size: 0.95rem;
      margin-top: 14px;
    }}
    @media (max-width: 1180px) {{
      .desktop-view {{
        display: none;
      }}
      .mobile-sheet {{
        display: block;
      }}
    }}
    @media (max-width: 900px) {{
      .page {{
        padding: 16px 10px 28px;
      }}
      .hero,
      .sheet-panel {{
        padding: 16px;
        border-radius: 16px;
      }}
      .tab {{
        flex: 1 1 calc(50% - 10px);
        text-align: center;
      }}
      .metric-grid {{
        grid-template-columns: 1fr;
      }}
      .contact-grid {{
        grid-template-columns: 1fr;
      }}
      .modal-head {{
        padding: 16px 16px 12px;
      }}
      .modal-media {{
        padding: 12px;
      }}
    }}
    @media (max-width: 540px) {{
      .page {{
        padding: 12px 8px 24px;
      }}
      .hero,
      .sheet-panel {{
        padding: 14px;
      }}
      .tab {{
        width: 100%;
      }}
      .mobile-card {{
        padding: 12px;
      }}
      .image-modal {{
        padding: 10px;
      }}
      .modal-card {{
        border-radius: 18px;
      }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <section class="hero">
      <h1>Lista de Precios Nu Skin Colombia 2025</h1>
      <p>Vista web generada directamente desde el Excel, respetando las categorias agrupadas, los colores por bloque y los enlaces oficiales de cada producto.</p>
      <div class="prompt-panel">
        <h2 class="prompt-panel-title">Datos para prompts de imagen</h2>
        <p class="prompt-help">Completa estos tres campos una sola vez. El emoji <strong>📋</strong> copiara un prompt listo para generar una imagen vertical para estados de WhatsApp usando la URL oficial del producto.</p>
        <div class="contact-grid">
          <label class="field">
            <span>Nombres</span>
            <input id="contact-name" type="text" placeholder="Tu nombre o marca personal">
          </label>
          <label class="field">
            <span>Celular</span>
            <input id="contact-cell" type="text" placeholder="Tu numero de contacto">
          </label>
          <label class="field">
            <span>No. Afiliado Marca</span>
            <input id="contact-affiliate" type="text" placeholder="Tu numero de afiliado o marca">
          </label>
        </div>
        <p id="clipboard-feedback" class="clipboard-feedback" aria-live="polite"></p>
      </div>
    </section>
    <div class="tabs">
      {"".join(tab_buttons)}
    </div>
    {"".join(sections)}
    <p class="footer">Archivo fuente: {escape(str(OUTPUT_XLSX))}</p>
  </main>
  <div id="image-modal" class="image-modal" hidden>
    <div class="modal-card" role="dialog" aria-modal="true" aria-labelledby="modal-title">
      <div class="modal-head">
        <div>
          <h3 id="modal-title" class="modal-title"></h3>
          <p id="modal-subtitle" class="modal-subtitle"></p>
        </div>
        <button type="button" class="modal-close" data-close-modal aria-label="Cerrar">×</button>
      </div>
      <div class="modal-media">
        <img id="modal-image" class="modal-image" alt="">
      </div>
    </div>
  </div>
  <div id="price-modal" class="image-modal" hidden>
    <div class="modal-card compact" role="dialog" aria-modal="true" aria-labelledby="price-modal-title">
      <div class="modal-head">
        <div>
          <h3 id="price-modal-title" class="modal-title"></h3>
          <p id="price-modal-subtitle" class="modal-subtitle"></p>
        </div>
        <button type="button" class="modal-close" data-close-modal aria-label="Cerrar">×</button>
      </div>
      <div id="price-modal-body" class="price-body"></div>
    </div>
  </div>
  <script>
    const tabs = document.querySelectorAll('.tab');
    const panels = document.querySelectorAll('.sheet-panel');
    const imageModal = document.getElementById('image-modal');
    const modalImage = document.getElementById('modal-image');
    const modalTitle = document.getElementById('modal-title');
    const modalSubtitle = document.getElementById('modal-subtitle');
    const priceModal = document.getElementById('price-modal');
    const priceModalTitle = document.getElementById('price-modal-title');
    const priceModalSubtitle = document.getElementById('price-modal-subtitle');
    const priceModalBody = document.getElementById('price-modal-body');
    const contactNameInput = document.getElementById('contact-name');
    const contactCellInput = document.getElementById('contact-cell');
    const contactAffiliateInput = document.getElementById('contact-affiliate');
    const clipboardFeedback = document.getElementById('clipboard-feedback');

    const contactFields = [
      {{ element: contactNameInput, key: 'nuskin-contact-name' }},
      {{ element: contactCellInput, key: 'nuskin-contact-cell' }},
      {{ element: contactAffiliateInput, key: 'nuskin-contact-affiliate' }},
    ];

    tabs.forEach((tab) => {{
      tab.addEventListener('click', () => {{
        const target = tab.dataset.target;
        tabs.forEach((item) => item.classList.toggle('is-active', item === tab));
        panels.forEach((panel) => panel.classList.toggle('is-active', panel.id === target));
      }});
    }});

    contactFields.forEach((field) => {{
      if (!field.element) {{
        return;
      }}
      try {{
        field.element.value = localStorage.getItem(field.key) || '';
      }} catch (error) {{
        field.element.value = '';
      }}
      field.element.addEventListener('input', () => {{
        try {{
          localStorage.setItem(field.key, field.element.value);
        }} catch (error) {{
          // Ignorar errores de almacenamiento local.
        }}
      }});
    }});

    function setClipboardFeedback(message, isError = false) {{
      if (!clipboardFeedback) {{
        return;
      }}
      clipboardFeedback.textContent = message;
      clipboardFeedback.style.color = isError ? '#b91c1c' : '#166534';
      if (!message) {{
        return;
      }}
      window.clearTimeout(setClipboardFeedback.timeoutId);
      setClipboardFeedback.timeoutId = window.setTimeout(() => {{
        if (clipboardFeedback.textContent === message) {{
          clipboardFeedback.textContent = '';
        }}
      }}, 2600);
    }}

    function getContactValue(input, fallback) {{
      return input && input.value.trim() ? input.value.trim() : fallback;
    }}

    function buildPromptText(button) {{
      const name = button.dataset.name || 'Producto Nu Skin';
      const sku = button.dataset.sku || '';
      const category = button.dataset.category || '';
      const officialUrl = button.dataset.url || '';
      const imageUrl = button.dataset.image || '';
      const contactName = getContactValue(contactNameInput, '[Completar nombres]');
      const contactCell = getContactValue(contactCellInput, '[Completar celular]');
      const contactAffiliate = getContactValue(contactAffiliateInput, '[Completar No. Afiliado Marca]');

      const lines = [
        `Crea una imagen publicitaria vertical 1080x1920 para compartir en estados de WhatsApp y promocionar el producto "${{name}}" de Nu Skin.`,
        sku ? `SKU: ${{sku}}.` : '',
        category ? `Categoria: ${{category}}.` : '',
        `Antes de diseñar la imagen, revisa obligatoriamente esta URL oficial de Nu Skin para extraer la informacion del producto y basar todo el contenido visual en esa fuente: ${{officialUrl}}`,
        'Usa solamente informacion verificable obtenida desde esa URL oficial.',
        'No inventes beneficios, no prometas resultados medicos y no uses claims no confirmados.',
        'La imagen debe verse premium, limpia, aspiracional y muy clara para venta por WhatsApp.',
        'La composicion debe incluir:',
        '1. El producto como protagonista visual.',
        '2. Un titular corto y fuerte basado en el beneficio principal oficial.',
        '3. Tres beneficios breves y verificables tomados de la URL oficial.',
        '4. Un llamado a la accion corto para pedir informacion o comprar.',
        '5. Un bloque final de contacto con exactamente estos datos:',
        `Nombres: ${{contactName}}`,
        `Celular: ${{contactCell}}`,
        `No. Afiliado Marca: ${{contactAffiliate}}`,
        imageUrl ? `Si necesitas referencia visual adicional del producto, usa esta imagen oficial como apoyo: ${{imageUrl}}` : '',
        'Entrega solo el prompt final para imagen, en español y sin explicaciones extra.',
      ];

      return lines.filter(Boolean).join('\\n');
    }}

    async function copyPromptToClipboard(button) {{
      const promptText = buildPromptText(button);
      try {{
        if (navigator.clipboard && window.isSecureContext) {{
          await navigator.clipboard.writeText(promptText);
        }} else {{
          const textArea = document.createElement('textarea');
          textArea.value = promptText;
          textArea.setAttribute('readonly', '');
          textArea.style.position = 'fixed';
          textArea.style.opacity = '0';
          document.body.appendChild(textArea);
          textArea.focus();
          textArea.select();
          document.execCommand('copy');
          document.body.removeChild(textArea);
        }}
        setClipboardFeedback(`Prompt copiado para ${{button.dataset.name || 'el producto'}}.`);
      }} catch (error) {{
        setClipboardFeedback('No se pudo copiar el prompt al portapapeles.', true);
      }}
    }}

    function buildPlaceholder(name, sku) {{
      const safeName = (name || 'Producto Nu Skin').replace(/[<>&]/g, '');
      const safeSku = (sku || '').replace(/[<>&]/g, '');
      const svg = `
        <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 960 960">
          <defs>
            <linearGradient id="bg" x1="0" y1="0" x2="1" y2="1">
              <stop offset="0%" stop-color="#edf5ff" />
              <stop offset="100%" stop-color="#fef6e7" />
            </linearGradient>
          </defs>
          <rect width="960" height="960" rx="42" fill="url(#bg)" />
          <rect x="140" y="160" width="680" height="470" rx="34" fill="#ffffff" stroke="#d9e5f3" stroke-width="10" />
          <circle cx="300" cy="330" r="76" fill="#dfeaf8" />
          <path d="M235 560l135-130 102 90 96-78 157 118" fill="none" stroke="#9fb7d3" stroke-width="26" stroke-linecap="round" stroke-linejoin="round" />
          <text x="120" y="730" fill="#1565c0" font-family="Segoe UI, Arial, sans-serif" font-size="44" font-weight="700">${{safeName}}</text>
          <text x="120" y="790" fill="#5f6b7a" font-family="Segoe UI, Arial, sans-serif" font-size="30">SKU ${{safeSku}}</text>
          <text x="120" y="848" fill="#5f6b7a" font-family="Segoe UI, Arial, sans-serif" font-size="28">Imagen oficial no disponible en la fuente actual.</text>
        </svg>`;
      return `data:image/svg+xml;charset=UTF-8,${{encodeURIComponent(svg)}}`;
    }}

    function escapeHtml(value) {{
      return String(value || '')
        .replace(/&/g, '&amp;')
        .replace(/</g, '&lt;')
        .replace(/>/g, '&gt;')
        .replace(/"/g, '&quot;')
        .replace(/'/g, '&#39;');
    }}

    function openImageModal(button) {{
      const name = button.dataset.name || 'Producto Nu Skin';
      const sku = button.dataset.sku || '';
      const imageUrl = button.dataset.image || buildPlaceholder(name, sku);
      modalTitle.textContent = name;
      modalSubtitle.textContent = sku ? `SKU ${{sku}}` : 'Producto Nu Skin';
      modalImage.src = imageUrl;
      modalImage.alt = name;
      imageModal.hidden = false;
      document.body.style.overflow = 'hidden';
    }}

    function openPriceModal(button) {{
      const name = button.dataset.name || 'Producto Nu Skin';
      const sku = button.dataset.sku || '';
      let prices = {{}};
      try {{
        prices = JSON.parse(button.dataset.prices || '{{}}');
      }} catch (error) {{
        prices = {{}};
      }}

      priceModalTitle.textContent = name;
      priceModalSubtitle.textContent = sku ? `SKU ${{sku}}` : 'Detalle de precios';
      priceModalBody.innerHTML = Object.entries(prices).map(([label, value]) => `
        <div class="price-row">
          <span>${{escapeHtml(label)}}</span>
          <strong>${{escapeHtml(value)}}</strong>
        </div>
      `).join('');
      priceModal.hidden = false;
      document.body.style.overflow = 'hidden';
    }}

    function closeImageModal() {{
      imageModal.hidden = true;
      modalImage.src = '';
      modalImage.alt = '';
      if (priceModal.hidden) {{
        document.body.style.overflow = '';
      }}
    }}

    function closePriceModal() {{
      priceModal.hidden = true;
      priceModalBody.innerHTML = '';
      if (imageModal.hidden) {{
        document.body.style.overflow = '';
      }}
    }}

    document.addEventListener('click', (event) => {{
      const emojiButton = event.target.closest('.img-emoji');
      if (emojiButton) {{
        openImageModal(emojiButton);
        return;
      }}

      const priceButton = event.target.closest('.price-emoji');
      if (priceButton) {{
        openPriceModal(priceButton);
        return;
      }}

      const promptButton = event.target.closest('.prompt-emoji');
      if (promptButton) {{
        copyPromptToClipboard(promptButton);
        return;
      }}

      if (event.target === imageModal) {{
        closeImageModal();
        return;
      }}

      if (event.target === priceModal) {{
        closePriceModal();
        return;
      }}

      if (event.target.closest('[data-close-modal]')) {{
        closeImageModal();
        closePriceModal();
      }}
    }});

    document.addEventListener('keydown', (event) => {{
      if (event.key === 'Escape') {{
        closeImageModal();
        closePriceModal();
      }}
    }});
  </script>
</body>
</html>
"""
    output_path.write_text(html, encoding="utf-8")


def safe_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number != number:
        return None
    return number


def format_number_local(value: float | None, decimals: int | None = None) -> str:
    if value is None:
        return "No disponible"
    if decimals is None:
        decimals = 0 if float(value).is_integer() else 1
    rendered = f"{value:,.{decimals}f}"
    return rendered.replace(",", "_").replace(".", ",").replace("_", ".")


def format_currency(value: float | None, decimals: int | None = None) -> str:
    if value is None:
        return "No disponible"
    return f"COP {format_number_local(value, decimals)}"


def format_detail_value(label: str, value: object) -> str:
    numeric = safe_float(value)
    if numeric is None:
        return "No disponible"
    if label == "VV":
        return format_number_local(numeric, 1)
    return format_currency(numeric)


def format_ratio_percent(value: float | None) -> str:
    if value is None:
        return "No disponible"
    return f"{format_number_local(value, 1)}%"


def format_date_es(dt: datetime) -> str:
    months = [
        "enero",
        "febrero",
        "marzo",
        "abril",
        "mayo",
        "junio",
        "julio",
        "agosto",
        "septiembre",
        "octubre",
        "noviembre",
        "diciembre",
    ]
    return f"{dt.day} de {months[dt.month - 1]} de {dt.year}"


def xml_escape(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def build_placeholder_data_uri(name: str, sku: str) -> str:
    label = xml_escape((name or "Producto Nu Skin")[:34])
    code = xml_escape(sku or "")
    svg = f"""
<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 960 960">
  <defs>
    <linearGradient id="bg" x1="0" y1="0" x2="1" y2="1">
      <stop offset="0%" stop-color="#edf5ff" />
      <stop offset="100%" stop-color="#fef6e7" />
    </linearGradient>
  </defs>
  <rect width="960" height="960" rx="42" fill="url(#bg)" />
  <rect x="140" y="160" width="680" height="470" rx="34" fill="#ffffff" stroke="#d9e5f3" stroke-width="10" />
  <circle cx="300" cy="330" r="76" fill="#dfeaf8" />
  <path d="M235 560l135-130 102 90 96-78 157 118" fill="none" stroke="#9fb7d3" stroke-width="26" stroke-linecap="round" stroke-linejoin="round" />
  <text x="120" y="730" fill="#1565c0" font-family="Segoe UI, Arial, sans-serif" font-size="44" font-weight="700">{label}</text>
  <text x="120" y="790" fill="#5f6b7a" font-family="Segoe UI, Arial, sans-serif" font-size="30">SKU {code}</text>
  <text x="120" y="848" fill="#5f6b7a" font-family="Segoe UI, Arial, sans-serif" font-size="28">Imagen oficial no disponible en la fuente actual.</text>
</svg>
""".strip()
    return f"data:image/svg+xml;charset=UTF-8,{quote(svg)}"


def is_generic_catalog_description(text: str) -> bool:
    sample = normalize_key(text)
    generic_prefixes = (
        "producto de la linea",
        "producto de higiene oral",
        "producto de cuidado",
        "producto o accesorio",
        "producto de la categoria",
        "producto de la linea pharmanex",
        "producto de la linea tru face",
        "producto de cuidado facial",
        "producto de cuidado personal",
    )
    return not sample or sample.startswith(generic_prefixes)


def pick_product_name(row: dict[str, object], metadata: dict[str, str]) -> str:
    row_name = clean_text(str(row.get("Descripcion del producto") or ""))
    candidates = [
        clean_text(metadata.get("title", "")),
        row_name,
        clean_text(metadata.get("name", "")),
    ]
    for candidate in candidates:
        if candidate and not candidate.isdigit():
            return candidate
    for candidate in candidates:
        if candidate:
            return candidate
    return f"SKU {row.get('SKU', '')}"


def sanitize_path_component(value: str, max_length: int = 100) -> str:
    value = clean_text(value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(char for char in value if not unicodedata.combining(char))
    value = re.sub(r"[^A-Za-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    if not value:
        value = "Producto"
    return value[:max_length].rstrip("_") or "Producto"


def ensure_product_record(
    product_metadata: dict[str, dict[str, str]],
    sku: str,
) -> dict[str, str]:
    return product_metadata.setdefault(
        sku,
        {
            "name": "",
            "description": "",
            "title": "",
            "category": "",
            "link": FALLBACK_PRODUCT_URL.format(sku=sku),
            "image_url": "",
            "thumbnail_url": "",
            "has_image": "false",
            "detail_page": "",
        },
    )


def build_detail_page_path(row: dict[str, object], product_metadata: dict[str, dict[str, str]]) -> Path:
    sku = str(row.get("SKU") or "").strip()
    metadata = ensure_product_record(product_metadata, sku)
    category = clean_text(str(row.get("Categoria") or metadata.get("category") or "Otros"))
    name = pick_product_name(row, metadata)
    folder = sanitize_path_component(category, max_length=60)
    filename = f"{sku}_{sanitize_path_component(name, max_length=110)}.html"
    return Path(folder) / filename


def assign_detail_page_paths(
    detail_rows: list[dict[str, object]],
    product_metadata: dict[str, dict[str, str]],
) -> dict[str, Path]:
    detail_map: dict[str, Path] = {}
    for row in detail_rows:
        sku = str(row.get("SKU") or "").strip()
        if not sku:
            continue
        output_path = build_detail_page_path(row, product_metadata)
        metadata = ensure_product_record(product_metadata, sku)
        metadata["detail_page"] = output_path.as_posix()
        detail_map[sku] = output_path
    return detail_map


def relative_href(from_path: Path, to_path: Path) -> str:
    return Path(os.path.relpath(to_path, start=from_path.parent)).as_posix()


def detect_product_mode(name: str, category: str) -> str:
    text = normalize_key(f"{name} {category}")
    if "suscrip" in text:
        return "subscription"
    if "kit" in text or "pack" in text or "pk" in text:
        return "kit"
    if any(keyword in text for keyword in ("cabezal", "cargador", "stand", "accent", "punta", "accesorio", "charger")):
        return "accessory"
    return "single"


def infer_profile_key(category: str, name: str) -> str:
    text = normalize_key(f"{category} {name}")
    category_key = normalize_key(category)
    if "ap-24" in category_key or any(keyword in text for keyword in ("tooth", "dental", "oral", "pasta")):
        return "oral"
    if any(keyword in text for keyword in ("galvanic", "lumispa", "device", "accent", "cabezal", "cargador", "stand")):
        return "device"
    if category_key == "nu colour" or any(keyword in text for keyword in ("lash", "brow", "lip", "color", "mascara")):
        return "beauty"
    if any(keyword in category_key for keyword in ("pharmanex", "tr90")) or any(keyword in text for keyword in ("collagen", "shake", "protein", "vitameal", "omega", "g3")):
        return "nutrition"
    if category_key == "cuidado corporal" or any(keyword in text for keyword in ("body", "foot", "hand", "lufra", "butter", "polishing bar")):
        return "bodycare"
    return "skincare"


def get_profile_copy(profile_key: str) -> dict[str, object]:
    profiles: dict[str, dict[str, object]] = {
        "oral": {
            "summary": "Este SKU se vende mejor como producto de uso diario, recompra alta y explicacion simple. La conversacion debe ir por experiencia de uso, frescura, rutina y margen, no por promesas exageradas.",
            "note": "Categoria facil de entender, demostracion corta y recompra natural.",
            "how_to_sell": [
                "Entrar por rutina diaria y sensacion de limpieza, no por tecnicismo.",
                "Usar comparacion simple contra opciones masivas sin pelear por precio.",
                "Cerrar con precio publico, precio miembro y link oficial.",
                "Hacer seguimiento corto para activar recompra o recomendacion.",
            ],
            "social_tips": [
                'WhatsApp: "Tengo una opcion premium para rutina diaria; si quieres te paso precio y link oficial".',
                "Mostrar una foto, un beneficio principal y el precio en tres lineas.",
                "Pedir respuesta cerrada: para ti, para tu pareja o para tu casa.",
                "Si es kit o pack, enfatizar ahorro por volumen y continuidad.",
            ],
        },
        "skincare": {
            "summary": "Este SKU entra por rutina, constancia y expectativa realista. En cuidado facial gana mas cuando el cliente entiende en que paso entra, que sensacion deja y como se complementa con otros productos.",
            "note": "No vender solo el frasco; vender el paso correcto dentro de la rutina.",
            "how_to_sell": [
                "Explicar si va en AM, PM o como tratamiento puntual.",
                "Usar antes y despues etico, prueba social y experiencia personal.",
                "Evitar promesas medicas; hablar de apariencia, sensacion y disciplina.",
                "Si es kit, vender comodidad y coherencia de rutina completa.",
            ],
            "social_tips": [
                "Story corta: problema visible, producto, CTA para recibir precio.",
                "Reel o video corto mostrando el paso dentro de la rutina.",
                "Responder objeciones con tiempo de uso y forma de aplicacion.",
                "Cerrar con pregunta concreta: quieres rutina basica o completa.",
            ],
        },
        "device": {
            "summary": "Los dispositivos, cabezales y accesorios se mueven mejor cuando se explican como complemento de un sistema existente. El valor comercial esta en compatibilidad, reposicion, mantenimiento y experiencia completa.",
            "note": "Si el cliente no entiende con que se usa, el cierre se cae.",
            "how_to_sell": [
                "Abrir la conversacion confirmando compatibilidad o equipo principal.",
                "Anclar el valor a reposicion, mantenimiento o extension de uso.",
                "Mostrar claramente si es accesorio, kit o equipo principal.",
                "Evitar venderlo aislado si el cliente aun no entiende el sistema.",
            ],
            "social_tips": [
                "Publicar foto o video junto al equipo compatible.",
                "Usar CTA tipo: necesitas reposicion o quieres completar tu set.",
                "Responder con SKU, compatibilidad y precio en el mismo mensaje.",
                "Si es suscripcion, enfatizar continuidad y simplificacion del pedido.",
            ],
        },
        "beauty": {
            "summary": "En belleza y color el cierre depende de resultado visible, facilidad de uso y conveniencia. Funciona bien con demostracion rapida, look final y prueba social breve.",
            "note": "Visual primero, explicacion despues.",
            "how_to_sell": [
                "Mostrar acabado, textura o efecto antes de hablar de precio.",
                "Usar mensajes cortos y muy visuales.",
                "Anclar el valor a practicidad, look y confianza al usarlo.",
                "Si es kit, remarcar que simplifica la compra y completa el look.",
            ],
            "social_tips": [
                "Story con antes, despues y CTA directo.",
                "Reel corto con aplicacion en tiempo real.",
                "Responder por DM con precio, disponibilidad y tono/uso.",
                "Invitar a prueba o demo si el formato lo permite.",
            ],
        },
        "bodycare": {
            "summary": "En cuidado corporal suele ganar la narrativa de problema concreto, sensorialidad y habito. El cliente entiende mejor la propuesta cuando se aterriza a resequedad, limpieza, suavidad o confort de la piel.",
            "note": "Problema concreto, uso simple y sensacion inmediata venden mejor.",
            "how_to_sell": [
                "Hablar del uso cotidiano y del beneficio mas tangible.",
                "Relacionar el producto con una molestia o necesidad clara.",
                "No saturar con datos; mostrar experiencia y resultado esperado.",
                "Si es barra, butter o tratamiento, explicar cuando se usa y cuanto rinde.",
            ],
            "social_tips": [
                "Foto de textura o aplicacion + CTA corto.",
                "Mensaje de WhatsApp centrado en una necesidad concreta.",
                "Seguimiento preguntando por sensacion o interes de rutina.",
                "Cruzar venta con otros productos de bienestar personal cuando aplique.",
            ],
        },
        "nutrition": {
            "summary": "En nutricion y bienestar el cierre depende de adherencia, acompañamiento y expectativa responsable. Se vende mejor como parte de un programa, un objetivo personal o una rutina sostenida.",
            "note": "Sin seguimiento no hay adherencia; sin adherencia no hay recompra.",
            "how_to_sell": [
                "Aterrizar el producto a un objetivo claro del cliente.",
                "Hablar de consumo, frecuencia y continuidad antes de cerrar.",
                "Evitar claims medicos o resultados absolutos.",
                "Si es kit o suscripcion, remarcar conveniencia y permanencia.",
            ],
            "social_tips": [
                "Usar historias de habito y constancia, no promesas extremas.",
                "Compartir forma de uso y tiempo estimado de consumo.",
                "Cerrar con pregunta de meta: energia, rutina, programa o bienestar.",
                "Programar seguimiento para sostener consumo y recompra.",
            ],
        },
    }
    return profiles.get(profile_key, profiles["skincare"])


def build_catalog_blurb(name: str, category: str, metadata: dict[str, str], profile_key: str, mode: str) -> str:
    description = clean_text(metadata.get("description", ""))
    if description and not is_generic_catalog_description(description):
        return description

    mode_copy = {
        "kit": "Se presenta como kit, asi que la venta funciona mejor cuando se explica el valor conjunto y la comodidad de comprar varios pasos a la vez.",
        "subscription": "Al estar en formato de suscripcion, conviene venderlo desde continuidad, recompra y simplificacion del pedido.",
        "accessory": "Al ser accesorio o reposicion, el argumento fuerte es compatibilidad, mantenimiento y extension del uso del sistema principal.",
        "single": "Su lectura comercial mejora cuando se aterriza a un beneficio claro, una rutina concreta y una objecion principal.",
    }
    family_copy = {
        "oral": "Pertenece a higiene oral y se mueve bien por rutina diaria, demostracion corta y recompra.",
        "skincare": "Pertenece a cuidado personal y requiere una explicacion clara del paso que ocupa dentro de la rutina.",
        "device": "Pertenece a una linea de dispositivos o complementos y se vende mejor unido al sistema correcto.",
        "beauty": "Pertenece a belleza y color, donde la prueba visual suele pesar mas que una descripcion larga.",
        "bodycare": "Pertenece a cuidado corporal y conviene llevarlo a una necesidad concreta de confort o apariencia.",
        "nutrition": "Pertenece a bienestar y nutricion, donde la continuidad y el acompanamiento son parte del cierre.",
    }
    return f"{name} forma parte de la categoria {category} de Nu Skin. {family_copy.get(profile_key, family_copy['skincare'])} {mode_copy.get(mode, mode_copy['single'])}"


def build_tags(category: str, mode: str, has_image: bool, public_price: float | None, member_price: float | None) -> list[str]:
    mode_labels = {
        "kit": "Formato kit",
        "subscription": "Suscripcion",
        "accessory": "Accesorio / reposicion",
        "single": "SKU individual",
    }
    tags = [
        category,
        mode_labels.get(mode, "SKU individual"),
        "Imagen oficial" if has_image else "Imagen referencial",
    ]
    if public_price is not None:
        tags.append(f"Publico {format_currency(public_price)}")
    if member_price is not None:
        tags.append(f"Miembro {format_currency(member_price)}")
    return tags


def bullet_list_html(items: list[str]) -> str:
    return "".join(f"<li>{escape(item)}</li>" for item in items)


def build_detail_page(
    row: dict[str, object],
    product_metadata: dict[str, dict[str, str]],
    generated_at: datetime,
    output_path: Path,
) -> str:
    sku = str(row.get("SKU") or "").strip()
    metadata = product_metadata.get(sku, {})
    category = clean_text(str(row.get("Categoria") or metadata.get("category") or "Otros"))
    name = pick_product_name(row, metadata)
    product_link = (metadata.get("link") or FALLBACK_PRODUCT_URL.format(sku=sku)).strip()
    profile_key = infer_profile_key(category, name)
    profile = get_profile_copy(profile_key)
    mode = detect_product_mode(name, category)
    blurb = build_catalog_blurb(name, category, metadata, profile_key, mode)

    public_price = safe_float(row.get("Precio Publico"))
    member_price = safe_float(row.get("Precio Miembro"))
    retail_bonus = safe_float(row.get("Bono minorista sin Impuesto"))
    vcv = safe_float(row.get("VCV"))
    vv = safe_float(row.get("VV"))
    spread = None
    markup = None
    if public_price is not None and member_price is not None:
        spread = public_price - member_price
        if member_price:
            markup = (spread / member_price) * 100

    image_url = (metadata.get("image_url") or metadata.get("thumbnail_url") or "").strip()
    image_url = image_url or build_placeholder_data_uri(name, sku)

    tags_html = "".join(f'<span class="tag">{escape(tag)}</span>' for tag in build_tags(category, mode, bool(metadata.get("image_url") or metadata.get("thumbnail_url")), public_price, member_price))

    detail_rows = []
    for label in PRICE_COLUMNS:
        detail_rows.append(
            "\n".join(
                [
                    "<tr>",
                    f"<td>{escape(label)}</td>",
                    f"<td>{escape(format_detail_value(label, row.get(label)))}</td>",
                    "</tr>",
                ]
            )
        )

    bonus_notes = {
        "4%": "Escalon base de referencia del bono por compartir o referir.",
        "8%": "Nivel intermedio calculado sobre el Precio Miembro sin Impuesto.",
        "12%": "Escalon medio dentro del esquema variable mostrado en la lista.",
        "16%": "Nivel alto de referencia para compartir y referir.",
        "20%": "Tope del bono por venta mostrado en el material LATAM 2025.",
        "24%": "Escenario maximo del bono por compartir de afiliado directo en la referencia visual.",
    }
    bonus_rows = []
    for label in BONUS_COLUMNS:
        bonus_rows.append(
            "\n".join(
                [
                    "<tr>",
                    f'<td><span class="legend-badge">{escape(label)}</span></td>',
                    f"<td>{escape(bonus_notes[label])}</td>",
                    f"<td>{escape(format_detail_value(label, row.get(label)))}</td>",
                    "</tr>",
                ]
            )
        )

    list_href = relative_href(output_path, OUTPUT_HTML)
    price_pdf_href = relative_href(output_path, OUTPUT_PRICE_PDF)
    sales_pdf_href = relative_href(output_path, OUTPUT_SALES_PLAN_PDF)

    sources = [
        f'<li><a href="{escape(product_link, quote=True)}" target="_blank" rel="noreferrer">Nu Skin Colombia · pagina oficial del SKU {escape(sku)}</a></li>',
        f'<li><a href="{escape(list_href, quote=True)}" target="_blank" rel="noreferrer">Lista web de precios 2025</a></li>',
        f'<li><a href="{escape(price_pdf_href, quote=True)}" target="_blank" rel="noreferrer">Lista detallada oficial Colombia 2025</a></li>',
        f'<li><a href="{escape(sales_pdf_href, quote=True)}" target="_blank" rel="noreferrer">Sales Performance Plan Overview LATAM 2025</a></li>',
    ]

    generated_note = (
        f"Pagina generada automaticamente el {format_date_es(generated_at)} "
        "a partir de la lista oficial 2025, el catalogo de Nu Skin Colombia y los metadatos disponibles del producto."
    )

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(sku)} | {escape(name)}</title>
  <style>
    :root {{
      --bg: #f6f8f7;
      --card: #ffffff;
      --text: #1d2a2d;
      --muted: #5e6b70;
      --line: #d8dedd;
      --brand: #0f766e;
      --accent: #d97706;
      --soft: #e8f6f3;
      --warn: #fff8e8;
      --shadow: 0 14px 34px rgba(20, 40, 44, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Arial, sans-serif;
      background:
        radial-gradient(circle at top left, rgba(15, 118, 110, 0.10), transparent 24%),
        linear-gradient(180deg, #fbfcfb 0%, var(--bg) 100%);
      color: var(--text);
    }}
    a {{ color: #0b62c4; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .page {{ max-width: 1120px; margin: 0 auto; padding: 24px 14px 48px; }}
    .hero, .card, .note {{
      background: var(--card);
      border: 1px solid rgba(29, 42, 45, 0.08);
      border-radius: 20px;
      box-shadow: var(--shadow);
    }}
    .hero {{
      display: grid;
      grid-template-columns: 220px 1fr;
      gap: 22px;
      padding: 24px;
      margin-bottom: 18px;
      align-items: center;
    }}
    .shot {{
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 250px;
      border-radius: 16px;
      border: 1px solid #d9ece7;
      background: linear-gradient(180deg, #f4fbfa 0%, #eef7f5 100%);
      padding: 12px;
    }}
    .shot img {{ width: 100%; max-width: 180px; height: auto; object-fit: contain; }}
    .eyebrow {{
      display: inline-block;
      padding: 8px 12px;
      border-radius: 999px;
      background: var(--soft);
      color: var(--brand);
      font-size: 0.82rem;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    h1 {{
      margin: 14px 0 10px;
      font-size: clamp(2rem, 4vw, 3rem);
      line-height: 1.03;
      color: #12423e;
    }}
    h2 {{
      margin: 0 0 14px;
      font-size: 1.35rem;
      color: #12423e;
    }}
    p, li {{
      line-height: 1.65;
      color: #334447;
    }}
    .subtitle {{ margin: 0 0 14px; color: var(--muted); }}
    .actions {{ display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 14px; }}
    .btn {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      padding: 11px 15px;
      border-radius: 12px;
      font-weight: 700;
      border: 1px solid transparent;
    }}
    .btn.primary {{ background: var(--brand); color: #fff; }}
    .btn.secondary {{ background: #fff7ed; border-color: #fed7aa; color: #9a4e07; }}
    .tags {{ display: flex; flex-wrap: wrap; gap: 8px; }}
    .tag {{
      padding: 8px 10px;
      border-radius: 999px;
      background: #f4f7f7;
      font-size: 0.9rem;
      color: #415255;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(12, minmax(0, 1fr));
      gap: 16px;
      margin-bottom: 16px;
    }}
    .card {{ padding: 20px; }}
    .col-12 {{ grid-column: span 12; }}
    .col-8 {{ grid-column: span 8; }}
    .col-6 {{ grid-column: span 6; }}
    .col-4 {{ grid-column: span 4; }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 12px;
      margin-top: 12px;
    }}
    .kpi {{
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 14px;
      background: #fbfcfc;
    }}
    .kpi small {{
      display: block;
      margin-bottom: 6px;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    .kpi strong {{
      display: block;
      font-size: 1.15rem;
      color: #12423e;
      margin-bottom: 6px;
    }}
    .note {{
      padding: 18px 20px;
      background: var(--warn);
      border-color: #fde6b6;
    }}
    .data-table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 0.94rem;
      margin-top: 10px;
    }}
    .data-table th,
    .data-table td {{
      border: 1px solid var(--line);
      padding: 10px 12px;
      text-align: left;
      vertical-align: top;
      line-height: 1.5;
    }}
    .data-table th {{ background: #f1f7f6; }}
    .legend-badge {{
      display: inline-flex;
      min-width: 58px;
      justify-content: center;
      padding: 7px 10px;
      border-radius: 999px;
      background: var(--soft);
      color: var(--brand);
      font-weight: 800;
    }}
    .source-list {{ margin: 0; padding-left: 18px; }}
    .footer {{
      margin-top: 18px;
      color: var(--muted);
      font-size: 0.92rem;
      line-height: 1.6;
      text-align: center;
    }}
    code {{
      background: #f3f5f5;
      padding: 2px 6px;
      border-radius: 6px;
      font-family: Consolas, "Courier New", monospace;
      font-size: 0.93em;
    }}
    @media (max-width: 980px) {{
      .hero {{ grid-template-columns: 1fr; }}
      .col-8, .col-6, .col-4 {{ grid-column: span 12; }}
      .kpis {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
    }}
    @media (max-width: 680px) {{
      .page {{ padding: 14px 10px 30px; }}
      .hero, .card, .note {{ padding: 16px; }}
      .data-table {{ display: block; overflow-x: auto; }}
      .kpis {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <section class="hero">
      <div class="shot">
        <img src="{escape(image_url, quote=True)}" alt="{escape(name)}">
      </div>
      <div>
        <span class="eyebrow">SKU {escape(sku)} · {escape(category)}</span>
        <h1>{escape(name)}</h1>
        <p class="subtitle">{escape(blurb)}</p>
        <div class="actions">
          <a class="btn primary" href="{escape(product_link, quote=True)}" target="_blank" rel="noreferrer">Producto oficial</a>
          <a class="btn secondary" href="{escape(list_href, quote=True)}">Volver a la lista</a>
          <a class="btn secondary" href="{escape(price_pdf_href, quote=True)}" target="_blank" rel="noreferrer">PDF 2025</a>
        </div>
        <div class="tags">
          {tags_html}
        </div>
        <div class="kpis">
          <div class="kpi">
            <small>Precio Público</small>
            <strong>{escape(format_currency(public_price))}</strong>
            <span>Precio de referencia al cliente final en la lista 2025.</span>
          </div>
          <div class="kpi">
            <small>Precio Miembro</small>
            <strong>{escape(format_currency(member_price))}</strong>
            <span>Base para compra interna, paquete o cierre relacional.</span>
          </div>
          <div class="kpi">
            <small>Spread Directo</small>
            <strong>{escape(format_currency(spread))}</strong>
            <span>Equivale a {escape(format_ratio_percent(markup))} de markup sobre el precio miembro.</span>
          </div>
          <div class="kpi">
            <small>Volumen</small>
            <strong>VCV {escape(format_currency(vcv))}</strong>
            <span>VV {escape(format_detail_value('VV', vv))} · Bono minorista {escape(format_currency(retail_bonus))}.</span>
          </div>
        </div>
      </div>
    </section>

    <section class="grid">
      <article class="card col-8">
        <h2>Resumen Comercial</h2>
        <p>{escape(str(profile["summary"]))}</p>
        <ul>
          {bullet_list_html(list(profile["how_to_sell"]))}
        </ul>
      </article>

      <aside class="note col-4">
        <strong>Lectura corta:</strong>
        <p>{escape(str(profile["note"]))}</p>
      </aside>
    </section>

    <section class="grid">
      <article class="card col-12">
        <h2>Convención Del Esquema De Ganancias</h2>
        <p>
          En la lista oficial de precios, las columnas <code>4%</code>, <code>8%</code>, <code>12%</code>, <code>16%</code>, <code>20%</code> y <code>24%</code>
          son montos de referencia para bonos por compartir y referir. Se leen junto con <code>Precio Miembro sin Impuesto</code>,
          <code>VCV</code>, <code>VV</code> y <code>Bono minorista sin Impuesto</code>.
        </p>
        <table class="data-table">
          <thead>
            <tr>
              <th>Porcentaje</th>
              <th>Qué representa</th>
              <th>Valor en este SKU</th>
            </tr>
          </thead>
          <tbody>
            {"".join(bonus_rows)}
          </tbody>
        </table>
      </article>
    </section>

    <section class="grid">
      <article class="card col-6">
        <h2>Detalle De Precios 2025</h2>
        <table class="data-table">
          <thead>
            <tr>
              <th>Campo</th>
              <th>Valor</th>
            </tr>
          </thead>
          <tbody>
            {"".join(detail_rows)}
          </tbody>
        </table>
      </article>

      <article class="card col-6">
        <h2>Promoción En WhatsApp Y Redes</h2>
        <ul>
          {bullet_list_html(list(profile["social_tips"]))}
        </ul>
      </article>
    </section>

    <section class="grid">
      <article class="card col-12">
        <h2>Fuentes Y Metodología</h2>
        <ul class="source-list">
          {"".join(sources)}
        </ul>
        <p>
          <strong>Nota metodológica:</strong> esta ficha resume la fila del SKU <strong>{escape(sku)}</strong> en la lista de precios 2025
          y añade una lectura comercial automatizada basada en la categoría, el formato del producto y sus precios publicados.
          No reemplaza la comunicación oficial de Nu Skin ni introduce claims médicos o de resultados garantizados.
        </p>
      </article>
    </section>

    <p class="footer">{escape(generated_note)}</p>
  </main>
</body>
</html>
"""
    return html


def export_detail_pages(
    detail_rows: list[dict[str, object]],
    product_metadata: dict[str, dict[str, str]],
    generated_at: datetime,
) -> set[Path]:
    written_paths: set[Path] = set()
    for row in detail_rows:
        sku = str(row.get("SKU") or "").strip()
        if not sku:
            continue
        output_href = detail_page_href(sku, product_metadata)
        output_path = Path(output_href) if output_href else build_detail_page_path(row, product_metadata)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        html = build_detail_page(row, product_metadata, generated_at, output_path)
        output_path.write_text(html, encoding="utf-8")
        written_paths.add(output_path)
    return written_paths


def cleanup_stale_detail_pages(valid_paths: set[Path]) -> int:
    valid = {path.as_posix() for path in valid_paths}
    removed = 0
    for candidate in Path(".").rglob("*.html"):
        relative = candidate.relative_to(Path("."))
        if not DETAIL_FILE_RE.fullmatch(candidate.stem):
            continue
        if relative.as_posix() in valid:
            continue
        candidate.unlink()
        removed += 1
    return removed


def style_sheet(ws, link_map: dict[str, str]) -> None:
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(color="FFFFFF", bold=True)
    category_palette = [
        "2E7D32",
        "1565C0",
        "7B1FA2",
        "EF6C00",
        "00838F",
        "5D4037",
        "AD1457",
        "6A1B9A",
        "283593",
        "0277BD",
    ]

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    numeric_columns = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
    for row in ws.iter_rows(min_row=2):
        for index in numeric_columns:
            row[index - 1].number_format = '#,##0.0'

    for row_idx in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=row_idx, column=2).value or "").strip()
        link = link_map.get(sku, "").strip()
        if not link:
            continue
        cell = ws.cell(row=row_idx, column=3)
        cell.hyperlink = link
        cell.style = "Hyperlink"
        cell.font = Font(color="0563C1", underline="single")

    categories = []
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if value and value not in categories:
            categories.append(value)
    color_map = {
        category: category_palette[index % len(category_palette)]
        for index, category in enumerate(categories)
    }

    start_row = 2
    while start_row <= ws.max_row:
        category = ws.cell(row=start_row, column=1).value
        end_row = start_row
        while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=1).value == category:
            end_row += 1

        fill = PatternFill("solid", fgColor=color_map.get(category, "607D8B"))
        for row_idx in range(start_row, end_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        start_row = end_row + 1

    autosize_sheet(ws)


def main() -> None:
    generated_at = datetime.now()
    text = PDF_TEXT.read_text(encoding="utf-8", errors="ignore")
    product_metadata = load_product_metadata()
    rows = extract_rows(text, product_metadata)

    if not rows:
        raise SystemExit("No se pudieron extraer filas del PDF.")

    df_all = pd.DataFrame(rows, columns=COLUMNS)
    df_all = df_all.sort_values(["Categoria", "SKU", "Descripcion del producto"], kind="stable").reset_index(drop=True)
    df_unique = df_all.drop_duplicates(subset=["SKU", "Descripcion del producto"]).copy()
    df_detail = df_all.drop_duplicates(subset=["SKU"], keep="first").copy()
    detail_rows = df_detail.to_dict(orient="records")
    price_list_skus = [sku for sku in df_all["SKU"].astype(str).unique() if sku]
    product_metadata = enrich_product_metadata_with_graphql(product_metadata, price_list_skus)
    assign_detail_page_paths(detail_rows, product_metadata)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="Todos", index=False)
        df_unique.to_excel(writer, sheet_name="Unicos", index=False)

    workbook = load_workbook(OUTPUT_XLSX)
    link_map = {
        sku: (metadata.get("link", "") or FALLBACK_PRODUCT_URL.format(sku=sku))
        for sku, metadata in product_metadata.items()
    }
    for sku in df_all["SKU"].astype(str).unique():
        link_map.setdefault(sku, FALLBACK_PRODUCT_URL.format(sku=sku))
    for sheet_name in workbook.sheetnames:
        style_sheet(workbook[sheet_name], link_map)
    workbook.save(OUTPUT_XLSX)
    export_workbook_html(workbook, OUTPUT_HTML, product_metadata)
    written_detail_paths = export_detail_pages(detail_rows, product_metadata, generated_at)
    removed_detail_pages = cleanup_stale_detail_pages(written_detail_paths)

    print(f"Filas extraidas: {len(df_all)}")
    print(f"Filas unicas: {len(df_unique)}")
    print(f"Fichas detalladas generadas: {len(written_detail_paths)}")
    print(f"Fichas detalladas limpiadas: {removed_detail_pages}")
    print(f"Excel: {OUTPUT_XLSX.resolve()}")
    print(f"HTML: {OUTPUT_HTML.resolve()}")


if __name__ == "__main__":
    main()
